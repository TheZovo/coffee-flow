from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models import Order, OrderItem, OrderStatus, Product
from app.schemas import (
    AdminAnalyticsDailyOut,
    AdminAnalyticsItemOut,
    AdminAnalyticsKpiOut,
    AdminAnalyticsOrderOut,
    AdminAnalyticsOut,
    AdminAnalyticsPromoOut,
    AdminAnalyticsStatusOut,
    AdminAnalyticsTopItemOut,
)


def default_period() -> tuple[date, date]:
    today = datetime.now().date()
    month_start = today.replace(day=1)
    return month_start, today


async def _load_orders(session: AsyncSession, date_from: date, date_to: date) -> list[Order]:
    result = await session.execute(
        select(Order)
        .options(
            joinedload(Order.user),
            joinedload(Order.promo_code),
            joinedload(Order.items).joinedload(OrderItem.product).joinedload(Product.category),
        )
        .where(Order.order_day >= date_from, Order.order_day <= date_to)
        .order_by(Order.created_at.desc(), Order.id.desc())
    )
    return result.unique().scalars().all()


def _completed_orders(orders: list[Order]) -> list[Order]:
    return [order for order in orders if order.status == OrderStatus.COMPLETED]


def _build_order_rows(orders: list[Order]) -> list[AdminAnalyticsOrderOut]:
    rows: list[AdminAnalyticsOrderOut] = []
    for order in orders:
        rows.append(
            AdminAnalyticsOrderOut(
                id=order.id,
                order_number=order.order_number,
                created_at=order.created_at,
                status=order.status.value,
                consumption_place=order.consumption_place,
                customer_name=order.contact_name,
                customer_phone=order.contact_phone,
                promo_code=order.promo_code.code if order.promo_code else None,
                total_cents=order.total_cents,
                discount_cents=order.discount_cents,
                final_cents=order.final_cents,
                bonus_spent_cents=order.bonus_spent_cents,
                bonus_earned_cents=order.bonus_earned_cents,
                loyalty_discount_cents=order.loyalty_discount_cents,
            )
        )
    return rows


def _build_item_rows(orders: list[Order]) -> list[AdminAnalyticsItemOut]:
    rows: list[AdminAnalyticsItemOut] = []
    for order in orders:
        for item in order.items:
            product = item.product
            category_slug = product.category.slug if product and product.category else None
            rows.append(
                AdminAnalyticsItemOut(
                    order_id=order.id,
                    order_number=order.order_number,
                    created_at=order.created_at,
                    status=order.status.value,
                    consumption_place=order.consumption_place,
                    customer_name=order.contact_name,
                    customer_phone=order.contact_phone,
                    promo_code=order.promo_code.code if order.promo_code else None,
                    product_id=item.product_id,
                    product_name=item.name_snapshot,
                    category_slug=category_slug,
                    qty=item.qty,
                    unit_price_cents=item.price_cents,
                    line_total_cents=item.price_cents * item.qty,
                    total_cents=order.total_cents,
                    discount_cents=order.discount_cents,
                    final_cents=order.final_cents,
                    bonus_spent_cents=order.bonus_spent_cents,
                    bonus_earned_cents=order.bonus_earned_cents,
                    loyalty_discount_cents=order.loyalty_discount_cents,
                )
            )
    return rows


def _top_items_by_product(orders: list[Order]) -> list[AdminAnalyticsTopItemOut]:
    totals: dict[str, dict[str, int | str]] = {}
    for order in _completed_orders(orders):
        for item in order.items:
            product = item.product
            key = str(item.product_id)
            current = totals.setdefault(
                key,
                {
                    "name": product.name if product else item.name_snapshot,
                    "qty": 0,
                    "revenue_cents": 0,
                },
            )
            current["qty"] = int(current["qty"]) + item.qty
            current["revenue_cents"] = int(current["revenue_cents"]) + item.qty * item.price_cents

    rows = [
        AdminAnalyticsTopItemOut(
            key=key,
            name=str(value["name"]),
            qty=int(value["qty"]),
            revenue_cents=int(value["revenue_cents"]),
        )
        for key, value in totals.items()
    ]
    return sorted(rows, key=lambda item: (-item.qty, -item.revenue_cents, item.name))


def _top_items_by_category(orders: list[Order]) -> list[AdminAnalyticsTopItemOut]:
    totals: dict[str, dict[str, int | str]] = {}
    for order in _completed_orders(orders):
        for item in order.items:
            product = item.product
            category_slug = product.category.slug if product and product.category else "uncategorized"
            category_name = product.category.name if product and product.category else "Без категории"
            current = totals.setdefault(
                category_slug,
                {
                    "name": category_name,
                    "qty": 0,
                    "revenue_cents": 0,
                },
            )
            current["qty"] = int(current["qty"]) + item.qty
            current["revenue_cents"] = int(current["revenue_cents"]) + item.qty * item.price_cents

    rows = [
        AdminAnalyticsTopItemOut(
            key=key,
            name=str(value["name"]),
            qty=int(value["qty"]),
            revenue_cents=int(value["revenue_cents"]),
        )
        for key, value in totals.items()
    ]
    return sorted(rows, key=lambda item: (-item.qty, -item.revenue_cents, item.name))


def _promo_usage(orders: list[Order]) -> list[AdminAnalyticsPromoOut]:
    usage: dict[str, dict[str, int]] = {}
    for order in _completed_orders(orders):
        code = order.promo_code.code if order.promo_code else None
        if not code:
            continue
        current = usage.setdefault(code, {"uses": 0, "discount_cents": 0})
        current["uses"] += 1
        current["discount_cents"] += order.promo_discount_cents
    rows = [
        AdminAnalyticsPromoOut(
            code=code,
            uses=value["uses"],
            discount_cents=value["discount_cents"],
        )
        for code, value in usage.items()
    ]
    return sorted(rows, key=lambda item: (-item.uses, -item.discount_cents, item.code))


def _status_breakdown(orders: list[Order]) -> list[AdminAnalyticsStatusOut]:
    counts = Counter(order.status.value for order in orders)
    return [
        AdminAnalyticsStatusOut(status=status, qty=qty)
        for status, qty in sorted(counts.items(), key=lambda item: item[0])
    ]


def _daily_revenue(orders: list[Order], date_from: date, date_to: date) -> list[AdminAnalyticsDailyOut]:
    counts = Counter(order.order_day for order in orders if order.order_day is not None)
    revenue = defaultdict(int)
    for order in _completed_orders(orders):
        if order.order_day is not None:
            revenue[order.order_day] += order.final_cents

    rows: list[AdminAnalyticsDailyOut] = []
    day = date_from
    while day <= date_to:
        rows.append(
            AdminAnalyticsDailyOut(
                day=day,
                orders=counts.get(day, 0),
                revenue_cents=revenue.get(day, 0),
            )
        )
        day += timedelta(days=1)
    return rows


def _build_kpis(orders: list[Order]) -> AdminAnalyticsKpiOut:
    completed = _completed_orders(orders)
    completed_count = len(completed)
    total_orders = len(orders)
    customer_counts = Counter(order.user_id for order in orders)
    completion_rate = round((completed_count / total_orders) * 100, 2) if total_orders else 0.0

    return AdminAnalyticsKpiOut(
        total_orders=total_orders,
        completed_orders=completed_count,
        cancelled_orders=sum(1 for order in orders if order.status == OrderStatus.CANCELLED),
        completion_rate=completion_rate,
        gross_revenue_cents=sum(order.total_cents for order in completed),
        net_revenue_cents=sum(order.final_cents for order in completed),
        average_order_value_cents=(sum(order.final_cents for order in completed) // completed_count) if completed_count else 0,
        unique_customers=len({order.user_id for order in orders}),
        repeat_customers=sum(1 for qty in customer_counts.values() if qty > 1),
        bonus_spent_cents=sum(order.bonus_spent_cents for order in completed),
        bonus_earned_cents=sum(order.bonus_earned_cents for order in completed),
        promo_discount_cents=sum(order.promo_discount_cents for order in completed),
        loyalty_discount_cents=sum(order.loyalty_discount_cents for order in completed),
    )


async def build_analytics(session: AsyncSession, date_from: date, date_to: date) -> AdminAnalyticsOut:
    orders = await _load_orders(session, date_from, date_to)
    return AdminAnalyticsOut(
        date_from=date_from,
        date_to=date_to,
        kpis=_build_kpis(orders),
        top_products=_top_items_by_product(orders),
        top_categories=_top_items_by_category(orders),
        promo_usage=_promo_usage(orders),
        status_breakdown=_status_breakdown(orders),
        daily_revenue=_daily_revenue(orders, date_from, date_to),
        orders=_build_order_rows(orders),
        items=_build_item_rows(orders),
    )


def build_analytics_csv_bytes(analytics: AdminAnalyticsOut) -> bytes:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "order_id",
            "order_number",
            "created_at",
            "status",
            "consumption_place",
            "customer_name",
            "customer_phone",
            "promo_code",
            "product_id",
            "product_name",
            "category_slug",
            "qty",
            "unit_price_cents",
            "line_total_cents",
            "order_total_cents",
            "discount_cents",
            "final_cents",
            "bonus_spent_cents",
            "bonus_earned_cents",
            "loyalty_discount_cents",
        ]
    )
    for item in analytics.items:
        writer.writerow(
            [
                item.order_id,
                item.order_number,
                item.created_at.isoformat(),
                item.status,
                item.consumption_place.value,
                item.customer_name or "",
                item.customer_phone or "",
                item.promo_code or "",
                item.product_id,
                item.product_name,
                item.category_slug or "",
                item.qty,
                item.unit_price_cents,
                item.line_total_cents,
                item.total_cents,
                item.discount_cents,
                item.final_cents,
                item.bonus_spent_cents,
                item.bonus_earned_cents,
                item.loyalty_discount_cents,
            ]
        )
    return output.getvalue().encode("utf-8-sig")


def build_analytics_xlsx_bytes(analytics: AdminAnalyticsOut) -> bytes:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in analytics.kpis.model_dump().items():
        summary_sheet.append([key, value])

    orders_sheet = workbook.create_sheet("Orders")
    orders_sheet.append(list(AdminAnalyticsOrderOut.model_fields.keys()))
    for row in analytics.orders:
        orders_sheet.append(
            [
                row.id,
                row.order_number,
                row.created_at.isoformat(),
                row.status,
                row.consumption_place.value,
                row.customer_name or "",
                row.customer_phone or "",
                row.promo_code or "",
                row.total_cents,
                row.discount_cents,
                row.final_cents,
                row.bonus_spent_cents,
                row.bonus_earned_cents,
                row.loyalty_discount_cents,
            ]
        )

    items_sheet = workbook.create_sheet("Items")
    items_sheet.append(list(AdminAnalyticsItemOut.model_fields.keys()))
    for row in analytics.items:
        items_sheet.append(
            [
                row.order_id,
                row.order_number,
                row.created_at.isoformat(),
                row.status,
                row.consumption_place.value,
                row.customer_name or "",
                row.customer_phone or "",
                row.promo_code or "",
                row.product_id,
                row.product_name,
                row.category_slug or "",
                row.qty,
                row.unit_price_cents,
                row.line_total_cents,
                row.total_cents,
                row.discount_cents,
                row.final_cents,
                row.bonus_spent_cents,
                row.bonus_earned_cents,
                row.loyalty_discount_cents,
            ]
        )

    top_products_sheet = workbook.create_sheet("TopProducts")
    top_products_sheet.append(["key", "name", "qty", "revenue_cents"])
    for row in analytics.top_products:
        top_products_sheet.append([row.key, row.name, row.qty, row.revenue_cents])

    top_categories_sheet = workbook.create_sheet("TopCategories")
    top_categories_sheet.append(["key", "name", "qty", "revenue_cents"])
    for row in analytics.top_categories:
        top_categories_sheet.append([row.key, row.name, row.qty, row.revenue_cents])

    promos_sheet = workbook.create_sheet("Promos")
    promos_sheet.append(["code", "uses", "discount_cents"])
    for row in analytics.promo_usage:
        promos_sheet.append([row.code, row.uses, row.discount_cents])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
